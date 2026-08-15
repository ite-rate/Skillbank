#!/usr/bin/env python3
"""Read and analyze Excel/CSV files with structured output.

v4 primary mode (workflow Step 1):
    python scripts/xlsx_reader.py file.xlsx --inspect

Legacy modes:
    python scripts/xlsx_reader.py file.xlsx               # basic preview
    python scripts/xlsx_reader.py file.xlsx --sheet Sales
    python scripts/xlsx_reader.py file.xlsx --stats
    python scripts/xlsx_reader.py file.xlsx --quality
    python scripts/xlsx_reader.py file.xlsx --json
"""

from __future__ import annotations

import sys
import json
import argparse
import re
import threading
import time
import zipfile
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError as e:
    print(json.dumps({
        "ok": False,
        "error_class": "missing_dependency",
        "human_summary": f"依赖缺失: {e}",
        "suggested_fix": {
            "action": "stop_and_report",
            "reason": "需要 pandas + openpyxl",
            "params": {"hint": "pip install pandas openpyxl"}
        }
    }, ensure_ascii=False))
    sys.exit(1)


# =========================================================================
# v4 INSPECT mode
# =========================================================================

CSV_ENCODINGS = ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']


def _skill_version():
    """读 SKILL 根目录 .version 文件的 version 字段. 仅打包后产物有此文件;
    本地开发返回 None. 调用方按 falsy 跳过, 避免污染输出. lru_cache 避免重读."""
    import functools as _ft
    if hasattr(_skill_version, '_cached'):
        return _skill_version._cached
    try:
        _v = (Path(__file__).resolve().parent.parent / '.version').read_text(
            encoding='utf-8', errors='replace')
        for _line in _v.splitlines():
            if _line.startswith('version:'):
                _skill_version._cached = _line.split(':', 1)[1].strip() or None
                return _skill_version._cached
    except Exception:
        pass
    _skill_version._cached = None
    return None

# Cap for _extract_formula_layer: stop scanning after this many cells
# to prevent OOM on large workbooks (100k+ rows × many columns).
MAX_FORMULA_SCAN_CELLS = 200_000

# INSPECT 采样上限：column_info / data_traps / formula-error 探测读取的最大数据行数。
# INSPECT 是快速陷阱探测步骤，不做全量分析（全量见 xlsx_analyze.py）。
# 20000 覆盖绝大多数业务表；超大文件另有 large_file trap 引导走 analyze。
INSPECT_SAMPLE_ROWS = 20000

# 大文件阈值（MB）：
#   > LARGE_FILE_WARN_MB → 加 large_file data_trap，引导后续 analyze 用 --head/--profile-only
#   > CSV_EXACT_COUNT_LIMIT_MB → CSV 跳过精确行数统计，改用字节估算（避免 30s+ 阻塞）
LARGE_FILE_WARN_MB = 100
CSV_EXACT_COUNT_LIMIT_MB = 500

# XLSX 流式 inspect 阈值（MB）：> 此值的 .xlsx/.xlsm 改走 openpyxl read_only=True 流式 +
# 单次解析路径，避免 load_workbook(read_only=False) 把整本灌内存。
# Windows 实测 42MB/30万行：read_only=False 峰值 3363MB/86s（>1分钟前台→被背景化→雪崩），
# read_only=True 仅 68MB/22s（省 49×）。低于阈值的常规文件保持原 read_only=False 全功能
# 路径（含原生合并单元格探测），零行为变化、零回归。
# 大文件流式路径的取舍（read_only 下经 Windows 实测）：
#   - ws.merged_cells 不可用 → 跳过合并单元格探测，输出诚实 hint + 按需 --merged-scan 命令
#   - ws.max_row 可能为 None → 行数标记为未精确统计（estimated）
#   - ws.cell()/number_format/data_type 仍可用 → 百分比/日期/公式错 trap 不受影响
#   - 列画像/预览直接从同一个 read_only wb 的 iter_rows 取（单次解析），不再二次 pd.read_excel
STREAMING_INSPECT_MB = 5

# inspect 扫描墙钟预算（秒）：formula 层 / formula 错误扫描超此预算主动停（返回已扫结果 +
# 截断标记）。跨平台用 time.monotonic（不依赖 Unix-only 的 SIGALRM；评测在 Windows 跑）。
# 防御性上限，与 MAX_FORMULA_SCAN_CELLS 双保险，确保任何单步都不顶满 1 分钟前台预算。
INSPECT_SCAN_BUDGET_SEC = 25.0

# 心跳：大文件 inspect 时，load_workbook / 扫描等都是「单个阻塞调用、中途不打印」，
# 而一些 Shell 执行环境的 idle timeout = 连续 60s 无 stdout/stderr 输出就切后台。
# 实测 42MB inspect 被切后台 2 次。对策：起一个后台线程每 HEARTBEAT_INTERVAL_SEC 向 stderr 打一行
# 进度，重置 idle 计时器，避免被切后台。心跳走 stderr（SKILL.md 标准命令不带 2>&1，
# stdout 仍是纯 JSON），且仅 > HEARTBEAT_MIN_MB 的文件才起（小文件 inspect 秒级，零影响）。
HEARTBEAT_MIN_MB = 3.0
HEARTBEAT_INTERVAL_SEC = 20.0

# inspect next_action 里给模型的极短脚本安全指针（完整 json-safe 模板在 SKILL.md，
# 这里只提醒，避免每次 inspect 输出膨胀）。命中模型「准备写分析脚本」那一刻。
_SCRIPT_SAFETY_HINT = (
    "写脚本: json.dump 必带 default= 防 int64/Timestamp 不可序列化(见 SKILL.md); "
    "建目录用结构化 file 工具非 shell mkdir; inspect/analyze 命令别加 2>&1"
)


class _Heartbeat:
    """后台线程，每隔 interval 秒向 stderr 打一行进度，防 Shell 命令 idle 切后台。"""
    def __init__(self, label: str, interval: float = HEARTBEAT_INTERVAL_SEC):
        self.label = label
        self.interval = interval
        self._stop = threading.Event()
        self._t = None
        self._start = None

    def start(self):
        self._start = time.monotonic()

        def _beat():
            while not self._stop.wait(self.interval):
                el = time.monotonic() - self._start
                try:
                    print(f'[inspect] {self.label}: 处理中 ~{el:.0f}s...',
                          file=sys.stderr, flush=True)
                except Exception:
                    return
        self._t = threading.Thread(target=_beat, daemon=True)
        self._t.start()
        return self

    def stop(self):
        self._stop.set()
        if self._t is not None:
            self._t.join(timeout=0.2)


def _estimate_csv_rows(filepath: str, encoding: str, file_size: int,
                        sample_rows: int = 1000) -> tuple[int, bool]:
    """Estimate CSV row count without scanning the whole file.

    Returns (count, is_estimated):
      - 若文件在 sample_rows 行内被扫完（小文件），返回精确行数 + is_estimated=False
      - 若达到 sample_rows 上限仍未到 EOF，按字节估算 + is_estimated=True
      - 空文件返回 (0, False)；异常返回 (-1, False)
    """
    try:
        with open(filepath, 'r', encoding=encoding) as f:
            head_bytes = 0
            actual_lines = 0
            hit_limit = False
            for line in f:
                head_bytes += len(line.encode(encoding, errors='replace'))
                actual_lines += 1
                if actual_lines >= sample_rows + 1:
                    hit_limit = True
                    break
            if actual_lines == 0:
                return 0, False
            if not hit_limit:
                # 全扫完了 — actual_lines 包含 header，返回精确行数
                return max(actual_lines - 1, 0), False
            avg_bytes = head_bytes / actual_lines
            if avg_bytes <= 0:
                return -1, False
            estimated = int(file_size / avg_bytes) - 1
        return max(estimated, 0), True
    except Exception:
        return -1, False

DYNAMIC_ARRAY_FUNCS = {
    'FILTER', 'SORT', 'SORTBY', 'UNIQUE', 'SEQUENCE', 'RANDARRAY',
    'XLOOKUP', 'XMATCH', 'LET', 'LAMBDA', 'MAP', 'REDUCE', 'SCAN',
    'BYROW', 'BYCOL', 'TEXTSPLIT', 'TEXTBEFORE', 'TEXTAFTER',
}


def _col_letter(idx: int) -> str:
    """1-based column index → A/B/.../AA/AB/..."""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def _sniff_csv_encoding(filepath: str) -> str | None:
    """Detect CSV encoding by sampling up to 256KB (not just 8KB)."""
    raw_sample = b''
    try:
        with open(filepath, 'rb') as f:
            raw_sample = f.read(256 * 1024)  # 256KB — covers most header+body
    except OSError:
        return None
    if not raw_sample:
        return 'utf-8'  # empty file — any encoding works
    for enc in CSV_ENCODINGS:
        try:
            raw_sample.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return None


def _check_vba_external(filepath: str) -> tuple[bool, bool, bool]:
    """Return (has_vba, has_external_links, has_pivot) via zip inspection."""
    has_vba = has_ext = has_pivot = False
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            names = zf.namelist()
            has_vba = any('vbaProject.bin' in n for n in names)
            has_ext = any('externalLink' in n for n in names)
            has_pivot = any('pivotTable' in n for n in names)
    except (zipfile.BadZipFile, FileNotFoundError):
        pass
    return has_vba, has_ext, has_pivot


def _extract_formula_layer(wb, deadline: float | None = None) -> dict:
    func_re = re.compile(r'([A-Z][A-Z0-9.]{1,})\s*\(')
    cross_re = re.compile(r"'?[^'!=\(\),\s]+?'?!")
    funcs: set[str] = set()
    cross_sheet = 0
    count = 0
    cells_scanned = 0
    truncated = False  # 触达 cell 上限或墙钟预算
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cells_scanned += 1
                if cells_scanned > MAX_FORMULA_SCAN_CELLS:
                    truncated = True
                    break
                # 每 5000 cell 查一次墙钟（避免每 cell 调 time 的开销）
                if (deadline is not None and cells_scanned % 5000 == 0
                        and time.monotonic() > deadline):
                    truncated = True
                    break
                if cell.data_type == 'f':
                    count += 1
                    f = str(cell.value or '')
                    for m in func_re.finditer(f):
                        funcs.add(m.group(1))
                    if cross_re.search(f):
                        cross_sheet += 1
            if truncated:
                break
        if truncated:
            break
    return {
        "count": count,
        "functions_used": sorted(funcs)[:30],
        "cross_sheet_refs": cross_sheet,
        "scan_truncated": truncated,
    }


# 句子标点(中英) — 表头列名不会含; 说明文字常含
_SENTENCE_PUNCT_RE = re.compile(r"[。？！；,，；：:][^\s\d]|http|www\.|@\w+\.|点此")
# 列名最大长度: 经验值 25 字符 (中文列名 8-12 字, 英文最多 30)
_HEADER_CELL_MAX_LEN = 25
# 列名最大长度方差比: 真表头列名长度相近 (例 [2,3,5,4,6]), 说明行常单格 100 字
_HEADER_LEN_RATIO = 3.0


def _looks_like_header_row(cells) -> bool:
    """启发式: 这一行像不像真表头? 不依赖关键词黑名单, 用语言学+结构特征."""
    if not cells:
        return False
    non_empty = []
    for c in cells:
        if c is None:
            continue
        s = str(c).strip()
        if s in ("", "nan", "None", "NaT"):
            continue
        non_empty.append(s)
    # 1. 列数: 表头通常 ≥ 3 列; 说明行/标题行通常 1-2 列
    if len(non_empty) < 3:
        return False
    # 2. 任一 cell 过长 → 是描述性句子, 不是列名
    if any(len(s) > _HEADER_CELL_MAX_LEN for s in non_empty):
        return False
    # 3. 任一 cell 含句子标点/URL/邮箱 → 是说明文字
    if any(_SENTENCE_PUNCT_RE.search(s) for s in non_empty):
        return False
    # 4. cell 长度差异: 列名长度通常接近, 不会一个 2 字一个 50 字
    lengths = [len(s) for s in non_empty]
    if len(lengths) > 1 and min(lengths) > 0:
        if max(lengths) / min(lengths) > _HEADER_LEN_RATIO and max(lengths) > 15:
            return False
    return True


def _row_looks_like_data(cells, header_cell_count=None) -> bool:
    """启发式: 这行是数据行吗?

    两个互补信号 (任一通过即视为数据):
      A. 含数字/日期/类型混合 (强信号)
      B. 与上一表头行列数对齐 (弱信号, 全字符串数据靠这个)
    """
    if not cells:
        return False
    non_empty = []
    types_seen = set()
    for c in cells:
        if c is None:
            continue
        if isinstance(c, (int, float)) and not (isinstance(c, float) and c != c):
            non_empty.append(c)
            types_seen.add("num")
            continue
        s = str(c).strip()
        if s in ("", "nan", "None", "NaT"):
            continue
        non_empty.append(s)
        try:
            float(s.replace(",", "").rstrip("%"))
            types_seen.add("numlike")
        except ValueError:
            if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", s):
                types_seen.add("date")
            else:
                types_seen.add("str")
    if len(non_empty) < 2:
        return False
    # A. 强信号: 含数字/日期或类型混合
    if ("num" in types_seen or "numlike" in types_seen
            or "date" in types_seen or len(types_seen) >= 2):
        return True
    # B. 弱信号: 全字符串但列数与表头对齐 (>= 60%)
    if header_cell_count and len(non_empty) >= header_cell_count * 0.6:
        return True
    return False


def _detect_skiprows_from_rows(raw_head_rows, max_scan: int = 10):
    """从 raw 前 N 行 (header=None 读取的 list of list) 推断真表头所在行 0-indexed.

    启发式 (不依赖关键词黑名单), 安全失败 (FN 优于 FP):
      A. 行像表头吗 (列数 ≥ 3 / 长度短且接近 / 无句子标点 / 无 URL)
      B. 必须能用下一行 (或容差下下一行) 验证是数据行, 否则不返回该 i
      C. 无下文可验证 → 返回 None (避免把"表头模板"误判)

    返回:
      - 0: row 0 就是真表头 (无需 skiprows)
      - N>0: 真表头在 row N, 用 skiprows=N
      - None: 推断不出 / 列数<3 / 无数据行可验证 (退化到模型看 previews.head)
    """
    if not raw_head_rows:
        return None
    rows = [list(r) if r is not None else [] for r in raw_head_rows[:max_scan]]
    for i, row in enumerate(rows):
        if not _looks_like_header_row(row):
            continue
        # 计算表头列数, 用于下一行的列对齐验证
        header_cells_n = sum(
            1 for c in row
            if c is not None and str(c).strip() not in ("", "nan", "None", "NaT")
        )
        # 必须能用下一行数据验证 — 没有下文则不接受 (避免 FP)
        if i + 1 < len(rows) and _row_looks_like_data(rows[i + 1], header_cells_n):
            return i
        if i + 2 < len(rows) and _row_looks_like_data(rows[i + 2], header_cells_n):
            return i
        # 此行像表头但无可验证数据 — 继续往下扫
    return None


def _detect_column_name_quirks(sheet_name, df, raw_head_rows=None) -> list[dict]:
    """检测列名中的"看不见"陷阱 — 模型自写 df['xxx'] 会 KeyError 的根因。

    陷阱来源（基于 1611 transcript 真实统计）：
      - leading/trailing 空白：' margin' / 'name ' → KeyError 'margin'/'name'
      - 半角括号前后空格：'库存成本 (元)' vs '库存成本(元)'（高频）
      - NBSP/全角空格/零宽空格：人眼看不出但 == 不成立
      - 换行/Tab：'地区\\n细分'（合并表头 ffill 副产物）
      - Unnamed: N：pandas 默认占位（合并 header 没正确处理）
      - 中文与 ASCII 间含空格：'系统 A' 实际是 '系统A'，模型反向加空格也会错

    单纯含半角括号（'金额(元)'）不报 — 那是合法列名。
    """
    if df is None or df.empty:
        return []
    quirks = []
    seen = set()  # 同 sheet 同 type 同 col 名去重
    for col_idx, col in enumerate(df.columns, start=1):
        name = str(col)
        col_letter = _col_letter(col_idx)
        pitfalls = []

        # 1. leading/trailing 空白
        if name != name.strip():
            pitfalls.append({
                "kind": "leading_or_trailing_whitespace",
                "detail": f"列名前/后含空白：{name!r}（strip 后 {name.strip()!r}）",
                "fix": "df = df.rename(columns=lambda c: c.strip())  # 或访问时也用 strip 一致",
            })

        # 2. 半角括号前/后多余空格："库存成本 (元)" / "金额( 元)"
        if re.search(r"\s\(|\(\s|\s\)|\)\s(?!$)", name):
            pitfalls.append({
                "kind": "space_around_paren",
                "detail": f"列名半角括号前后含空格：{name!r}（与无空格版本不等价）",
                "fix": (
                    "# 1) 严格按字面访问：df[{n!r}]\n"
                    "# 2) 或规范化：df = df.rename(columns=lambda c: re.sub(r'\\\\s+\\\\(', '(', re.sub(r'\\\\(\\\\s+', '(', c)))"
                ).format(n=name),
            })

        # 3. NBSP / 全角空格 / 零宽空格
        if any(ch in name for ch in (" ", "　", "​", "‌", "﻿")):
            chars_found = [f"U+{ord(c):04X}" for c in name if c in " 　​‌﻿"]
            pitfalls.append({
                "kind": "invisible_unicode",
                "detail": f"列名含不可见 Unicode（{','.join(set(chars_found))}）：{name!r}",
                "fix": (
                    "import unicodedata\n"
                    "df.columns = [unicodedata.normalize('NFKC', c).replace('\\u200b','').strip() for c in df.columns]"
                ),
            })

        # 4. 换行 / Tab
        if "\n" in name or "\t" in name or "\r" in name:
            pitfalls.append({
                "kind": "newline_or_tab",
                "detail": f"列名含换行/Tab：{name!r}（合并表头 ffill 副产物，常见多级表头）",
                "fix": (
                    "# 多级表头建议用 header=[0,1] 读；如已读成单层，规范化：\n"
                    "df.columns = [re.sub(r'[\\\\s\\\\n\\\\r\\\\t]+', '_', c).strip('_') for c in df.columns]"
                ),
            })

        # 5. Unnamed: N（pandas 默认占位 — 合并表头/前缀说明行未处理）
        if re.match(r"^Unnamed:\s*\d+$", name):
            # 内部启发式仅给"可能性提示", 不强推具体 N (避免多级表头/短词标签场景误导)
            hint = _detect_skiprows_from_rows(raw_head_rows) if raw_head_rows else None
            if hint is not None and hint > 0:
                hint_line = f"# 启发式推测前 {hint} 行是说明/标签 (看 previews.head 验证后再用)"
            else:
                hint_line = "# 请查看 previews.head 前几行原始数据, 二选一:"
            fix_text = (
                f"{hint_line}\n"
                f"# 方案 A · 前缀说明/空行场景: df = pd.read_excel('<file>', sheet_name={sheet_name!r}, skiprows=N)\n"
                f"# 方案 B · 多级表头场景:    df = pd.read_excel('<file>', sheet_name={sheet_name!r}, header=[0,1])\n"
                f"# 必须看 previews.head 实际内容决定 (启发式无法区分'短词标签行+表头' vs '多级表头')"
            )
            pitfalls.append({
                "kind": "unnamed_placeholder",
                "detail": f"列名为 pandas 默认占位 {name!r}（合并表头/前缀说明行 未跳过）",
                "fix": fix_text,
                "detected_skiprows_hint": hint,
            })

        # 6. 中文字与 ASCII 字母/数字间的空格（'系统 A' / '用户 ID'）
        if re.search(r"[一-鿿]\s+[A-Za-z0-9]|[A-Za-z0-9]\s+[一-鿿]", name):
            pitfalls.append({
                "kind": "cjk_ascii_space",
                "detail": f"列名中文与字母/数字间含空格：{name!r}（容易被模型自动'去空格'误访问）",
                "fix": f"# 严格按字面访问 df[{name!r}]；脚本里禁止"
                       "自行 strip 中间空格",
            })

        for p in pitfalls:
            key = (sheet_name, name, p["kind"])
            if key in seen:
                continue
            seen.add(key)
            entry = {
                "type": "column_name_quirk",
                "sheet": sheet_name,
                "col": col_letter,
                "name": name,
                "name_repr": repr(name),
                "pitfall": p["kind"],
                "note": p["detail"],
                "suggested_fix_code": p["fix"],
            }
            if "detected_skiprows_hint" in p and p["detected_skiprows_hint"] is not None:
                entry["detected_skiprows_hint"] = p["detected_skiprows_hint"]
            quirks.append(entry)

    return quirks


def _recommend_validate_columns(column_info: list[dict]) -> dict | None:
    """schema 是否含『模型容易反向写错』的列名特征 — 含则建议预校验。

    基于 evalrun_5a8685ac7450（33 trial）实证：
      - 模型在合法 schema 列名上反向加空格/改英文 alias 的高发场景
      - "库存成本(元)" → 模型写 "库存成本 (元)"  → KeyError
      - "客户ID"       → 模型写 "客户 ID"        → KeyError
      - "下单日期"     → 模型写 "下单_date"      → KeyError
    条件触发避免对干净 schema (24/33 case) 多增 tool 调用。

    Args:
      column_info: inspect 收集的所有列描述。

    Returns:
      None — 没有可疑特征，不增 hint
      dict — { reason, sample_columns, hint }，调用方拼到 next_action 里
    """
    if not column_info:
        return None
    risky_cols: list[str] = []
    seen_reasons: set[str] = set()
    for c in column_info:
        name = str(c.get("name", ""))
        if not name:
            continue
        # 1) 半角括号（"库存成本(元)" / "解决时长(小时)"）
        if re.search(r"\([^()]*\)", name):
            risky_cols.append(name)
            seen_reasons.add("paren")
            continue
        # 2) 中英混排（"客户ID" / "用户ID" / "SKU编号"）
        if re.search(r"[一-鿿][A-Za-z0-9]|[A-Za-z0-9][一-鿿]", name):
            risky_cols.append(name)
            seen_reasons.add("cjk_ascii_mix")
            continue
        # 3) 中文 + 下划线（"销售额_目标" — 模型常 alias 成英文）
        if "_" in name and re.search(r"[一-鿿]", name):
            risky_cols.append(name)
            seen_reasons.add("cjk_underscore")
            continue
    if not risky_cols:
        return None
    # 限制 hint 长度：最多列 3 个 sample
    samples = risky_cols[:3]
    reasons_zh = {
        "paren": "含半角括号",
        "cjk_ascii_mix": "中英混排",
        "cjk_underscore": "中文+下划线",
    }
    reason_label = " / ".join(reasons_zh[r] for r in seen_reasons)
    return {
        "risky_count": len(risky_cols),
        "sample_columns": samples,
        "reason": reason_label,
        "hint": (
            f"⚠️ schema 有 {len(risky_cols)} 个列名含『{reason_label}』特征（如："
            + "、".join(f"'{s}'" for s in samples)
            + "），模型在脚本里反向加空格/改英文别名极易触发 KeyError。"
            + "写完脚本首次运行前，建议把脚本里 df['xxx'] 用到的列名跑一次预校验：\n"
            + "  python scripts/xlsx_reader.py <file> --validate-columns 列1 列2 ..."
        ),
    }


def _detect_duplicate_headers(sheet_name, ws) -> list[dict]:
    """检测 sheet 表头行中的重复列名。

    pandas.read_excel 遇到重复列名时静默把第 2 个起重命名为 `name.1`/`name.2`，
    模型对此无感知，可能把两列当一列分析或在 KPI 中混淆口径。
    """
    from collections import Counter
    traps: list[dict] = []
    try:
        first_row = next(ws.iter_rows(max_row=1, values_only=True))
    except StopIteration:
        return traps
    names = [str(v).strip() if v is not None else "" for v in first_row]
    # 忽略空字符串（合并表头/前缀说明行的占位）
    counts = Counter(n for n in names if n)
    dups = {n: c for n, c in counts.items() if c >= 2}
    if dups:
        # 同名列在 raw header 中的位置（A/B/C... 字母）
        positions = {}
        for n in dups:
            positions[n] = [_col_letter(i + 1) for i, v in enumerate(names) if v == n]
        traps.append({
            "type": "duplicate_header",
            "sheet": sheet_name,
            "duplicates": dups,
            "positions": positions,
            "note": (
                f"sheet '{sheet_name}' 表头有 {len(dups)} 组重复列名："
                + "、".join(f"「{n}」×{c}（{','.join(positions[n])}）"
                           for n, c in dups.items())
                + "；pandas.read_excel 会静默把第 2 个起重命名为 'name.1'，"
                  "若 KPI 仅引用『销售额』可能口径混淆"
            ),
            "suggested_read_code": (
                "# 重复列名场景：读完后必须显式 rename 区分口径\n"
                "import pandas as pd\n"
                f"df = pd.read_excel('<file>', sheet_name='{sheet_name}')\n"
                f"# 默认重命名后类似：{list(dups.keys())[0]}, {list(dups.keys())[0]}.1, ...\n"
                "# 必须基于业务含义改名：\n"
                "# df = df.rename(columns={'X.1': 'X_含税', 'X': 'X_不含税'})  # 示例\n"
                "# 然后在 KPI label 中显式标注每个口径"
            ),
        })
    return traps


def _profile_scope_entry(rows_total: int, rows_profiled: int) -> dict:
    """构造单 sheet 的画像采样范围说明（profile_scope 字段值）。

    rows_total    = sheet 真实数据行数（不含表头）
    rows_profiled = INSPECT 实际读入做画像/陷阱探测的数据行数
    """
    sampled = rows_total > rows_profiled
    if sampled:
        note = (
            f"列画像（dtype/null率/samples）与陷阱探测仅覆盖前 {rows_profiled} 行；"
            f"第 {rows_profiled + 1} 行起共 {rows_total - rows_profiled} 行未扫描，"
            f"如需全量统计请用 xlsx_analyze.py"
        )
    else:
        note = f"全表 {rows_total} 行已全部纳入画像"
    return {
        "rows_total": rows_total,
        "rows_profiled": rows_profiled,
        "sampled": sampled,
        "note": note,
    }


def _inspect_sampled_trap(sheet_name: str, rows_total: int,
                          rows_profiled: int) -> dict:
    """sheet 被采样时追加到 data_traps 的说明条目（确保模型注意到画像非全表）。"""
    return {
        "type": "inspect_sampled",
        "sheet": sheet_name,
        "rows_profiled": rows_profiled,
        "rows_total": rows_total,
        "note": (
            f"sheet '{sheet_name}' 共 {rows_total} 行，INSPECT 仅采样前 {rows_profiled} 行"
            f"做 column_info 与 data_traps 探测；第 {rows_profiled + 1} 行后的脏数据/"
            f"合并单元格/公式错不在本清单内。previews.tail 已单独取真实表尾。"
        ),
        "suggested_read_code": (
            "# 列画像仅基于前 N 行采样，全量统计请用 analyze：\n"
            "# python xlsx_analyze.py '<file>'\n"
            "# 读全量数据（不要加 nrows）：\n"
            f"# df = pd.read_excel('<file>', sheet_name='{sheet_name}')"
        ),
    }


def _real_tail_xlsx(ws, ncols: int, n: int = 5) -> list:
    """用 openpyxl 直接读 sheet 的真实表尾 n 个数据行。

    被截断采样时 df.tail() 只能拿到采样段尾巴，取不到真实表尾（签字栏/合计行）。
    """
    max_row = int(ws.max_row or 0)
    if max_row <= 1 or ncols <= 0:
        return []
    start = max(2, max_row - n + 1)
    rows = []
    for r in ws.iter_rows(min_row=start, max_row=max_row,
                          max_col=ncols, values_only=True):
        rows.append(['' if v is None else str(v) for v in r])
    return rows


def _dedupe_columns(header: list[str]) -> list[str]:
    """模拟 pandas.read_excel 的列名处理：空 cell → 'Unnamed: i'，
    重复列名 → 'name'、'name.1'、'name.2'…。供大文件流式路径构造 df 用，
    使 column_info / _detect_column_name_quirks 行为与 read_excel 路径一致。"""
    counts: dict[str, int] = {}
    out: list[str] = []
    for i, h in enumerate(header):
        base = h if h not in ('', 'None', 'nan', 'NaT') else f'Unnamed: {i}'
        if base in counts:
            counts[base] += 1
            out.append(f'{base}.{counts[base]}')
        else:
            counts[base] = 0
            out.append(base)
    return out


def _sample_df_via_openpyxl(ws, max_rows: int):
    """大文件流式路径：从已打开的 read_only worksheet 用 iter_rows 取前 max_rows 行
    构造 DataFrame —— 复用 wb 已载入的 shared strings，避免二次 pd.read_excel 重新
    解析整本（Windows 42MB 实测二次解析 ~26s）。首行作表头。返回 (df, rows_read)。"""
    import pandas as pd
    header = None
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows + 1,
                                         values_only=True)):
        if i == 0:
            header = ['' if v is None else str(v) for v in row]
            continue
        rows.append(list(row))
    if not header:
        return pd.DataFrame(), 0
    df = pd.DataFrame(rows, columns=_dedupe_columns(header))
    # 让 pandas 推断每列 dtype（与 read_excel 接近：数字列转数值、其余 object）
    try:
        df = df.infer_objects()
    except Exception:
        pass
    return df, len(rows)


def _iterparse_merged_cells(filepath: str, max_per_sheet: int = 20,
                            deadline: float | None = None) -> dict:
    """从 xlsx 的 sheet XML 用 iterparse 流式取合并单元格区域（低内存，不全量载入）。
    供 --merged-scan 按需调用；大文件流式 inspect 默认不跑（Windows 42MB 实测 ~23s）。
    返回 {sheet_xml: [refs...]}（sheet_xml 顺序近似 workbook 顺序，best-effort）。"""
    import zipfile
    import xml.etree.ElementTree as ET
    NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    out: dict = {}

    def _sheet_num(n):  # 按文件名里的数字排序（避免 sheet10 排到 sheet2 前）
        m = re.search(r'sheet(\d+)\.xml$', n)
        return int(m.group(1)) if m else 1 << 30

    with zipfile.ZipFile(filepath) as z:
        sheet_xmls = sorted(
            (n for n in z.namelist()
             if re.search(r'xl/worksheets/sheet\d+\.xml$', n)),
            key=_sheet_num)
        for sx in sheet_xmls:
            refs: list[str] = []
            seen = 0
            with z.open(sx) as f:
                for _ev, elem in ET.iterparse(f, events=('end',)):
                    tag = elem.tag
                    if tag == NS + 'mergeCell':
                        ref = elem.get('ref')
                        if ref and len(refs) < max_per_sheet:
                            refs.append(ref)
                        elem.clear()
                    elif tag == NS + 'c' or tag == NS + 'row':
                        elem.clear()
                        seen += 1
                        if (deadline is not None and seen % 50000 == 0
                                and time.monotonic() > deadline):
                            break
            out[Path(sx).name] = refs
    return out


def _detect_formula_errors(sheet_name, ws,
                           max_rows: int = INSPECT_SAMPLE_ROWS,
                           deadline: float | None = None) -> list[dict]:
    """检测 sheet 中的公式错误值（#DIV/0!/#VALUE!/#REF!/#NAME?/#N/A/#NULL!/#NUM!）。

    pandas.read_excel 会把这些错误值转成 NaN，模型若直接用 .fillna(0) 或忽略
    NaN 会得到错误聚合结果。本检测让模型知道 NaN 来源是公式失败而非数据缺失。

    扫描范围 = 数据行 2..max_rows（与 inspect 采样上限 INSPECT_SAMPLE_ROWS 对齐），
    细胞数额外受 MAX_FORMULA_SCAN_CELLS=200k 约束。两个上限都会在 trap 输出中告知。
    """
    traps: list[dict] = []
    by_col: dict[str, dict] = {}
    cells_scanned = 0
    last_row_seen = 1
    cell_cap_hit = False
    try:
        for row in ws.iter_rows(min_row=2, max_row=max_rows, values_only=False):
            for cell in row:
                cells_scanned += 1
                last_row_seen = max(last_row_seen, cell.row)
                if cells_scanned > MAX_FORMULA_SCAN_CELLS:
                    cell_cap_hit = True
                    break
                if (deadline is not None and cells_scanned % 5000 == 0
                        and time.monotonic() > deadline):
                    cell_cap_hit = True
                    break
                if cell.data_type == 'e':
                    letter = cell.column_letter
                    if letter not in by_col:
                        by_col[letter] = {
                            "col": letter,
                            "error_values": [],
                            "first_row": cell.row,
                            "count": 0,
                        }
                    by_col[letter]["count"] += 1
                    if len(by_col[letter]["error_values"]) < 3:
                        by_col[letter]["error_values"].append(str(cell.value))
            if cell_cap_hit:
                break
    except Exception:
        return traps

    # ws.max_row may be much larger than max_rows; tell model what wasn't seen
    try:
        sheet_max_row = int(ws.max_row or 0)
    except Exception:
        sheet_max_row = last_row_seen
    rows_unscanned = max(0, sheet_max_row - max_rows)

    if by_col:
        total = sum(b["count"] for b in by_col.values())
        note_extra = ""
        if rows_unscanned > 0:
            note_extra = (f"；本扫描仅覆盖前 {max_rows} 行，"
                          f"sheet 共 {sheet_max_row} 行，"
                          f"后 {rows_unscanned} 行未扫，可能有更多公式错误")
        if cell_cap_hit:
            note_extra += f"；细胞数达 {MAX_FORMULA_SCAN_CELLS} 上限，扫描提前终止"
        traps.append({
            "type": "formula_error_cells",
            "sheet": sheet_name,
            "total_error_cells": total,
            "by_column": list(by_col.values()),
            "rows_scanned": min(sheet_max_row, max_rows),
            "rows_in_sheet": sheet_max_row,
            "rows_unscanned": rows_unscanned,
            "scanned_capped": cell_cap_hit,
            "note": (
                f"sheet '{sheet_name}' 前 {min(sheet_max_row, max_rows)} 行检出 "
                f"{total} 个公式错误单元格（"
                + "、".join(f"{b['col']}列×{b['count']}（首例 {b['error_values'][0]}）"
                           for b in by_col.values())
                + "）；pandas.read_excel 会把这些位置转为 NaN，"
                  "直接 sum/mean 会忽略而非告警，建议先 .isna().sum() 报告失败行数"
                + note_extra
            ),
            "suggested_read_code": (
                "# 公式错误值场景：读完后报告失败行而非默默忽略\n"
                "# pandas 会把 #DIV/0!/#VALUE!/#REF! 等转成 NaN，与『真实缺失』无法直接区分。\n"
                "\n"
                "# ───── 第一段：纯 pandas 处理（立即可跑）─────\n"
                "import pandas as pd\n"
                f"df = pd.read_excel('<file>', sheet_name='{sheet_name}')\n"
                "# 1) 聚合前先统计 NaN 行数\n"
                "err_rows = df[df['<col>'].isna()]\n"
                "print(f'<col> NaN 行数: {len(err_rows)}, 占比: {len(err_rows)/len(df)*100:.1f}%')\n"
                "# 2) 若 NaN 比例 > 5%，必须在报告中显式标注\n"
                "# 3) 聚合时显式 .dropna() 而非 .fillna(0)，并标注 n_excluded\n"
                f"agg = df['<col>'].dropna().sum()\n"
                "print(f'<col> sum (excluding NaN): {agg}, n_excluded: {len(err_rows)}')\n"
                "\n"
                "# ───── 第二段：如需区分『公式失败』vs『真实缺失』─────\n"
                "# pandas DataFrame 不保留 cell.data_type 信息；需用 openpyxl 单独取错误位置。\n"
                "# 本 trap 的 `by_column` 字段已列出错误位置（行号 + 错误值），直接用即可：\n"
                "#   formula_err_cells = [(b['col'], b['first_row'], b['error_values'])\n"
                "#                        for b in trap['by_column']]\n"
                "# 或自己 openpyxl 扫一遍：\n"
                "#   import openpyxl\n"
                "#   wb = openpyxl.load_workbook('<file>', data_only=True)\n"
                "#   ws = wb['" + sheet_name + "']\n"
                "#   err_coords = [c.coordinate for row in ws.iter_rows() for c in row if c.data_type == 'e']"
            ),
        })
    return traps


def _detect_sheet_traps(sheet_name, df, ws, streaming: bool = False) -> list[dict]:
    traps: list[dict] = []

    # 大文件流式（read_only）下 ws.merged_cells 不可用（AttributeError）；
    # 跳过合并单元格探测（调用方另发诚实 hint + --merged-scan 命令）。
    # 注意：number_format / ws.cell() 在 read_only 下仍可用，下面的百分比/日期 trap 照常跑。
    if not streaming and ws.merged_cells and ws.merged_cells.ranges:
        ranges = list(ws.merged_cells.ranges)
        for mr in ranges[:5]:
            traps.append({
                "type": "merged_cell",
                "sheet": sheet_name,
                "range": str(mr),
                "note": "pandas.read_excel 会丢失合并单元格的非左上角值",
                "suggested_read_code": (
                    f'# 多级表头场景：用 header=[0,1]\n'
                    f'df = pd.read_excel("<file>", sheet_name="{sheet_name}", header=[0,1])\n'
                    f'# 或者合并值延续下填：\n'
                    f'df = pd.read_excel("<file>", sheet_name="{sheet_name}")\n'
                    f'df = df.ffill()  # 沿列延续合并单元格的顶部值'
                ),
            })
        if len(ranges) > 5:
            traps.append({
                "type": "merged_cell",
                "sheet": sheet_name,
                "range": f"(another {len(ranges) - 5} merged ranges)",
                "note": "更多合并单元格未列出",
            })

    if df is None or df.empty:
        return traps

    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = _col_letter(col_idx)
        try:
            cell = ws.cell(row=2, column=col_idx)
        except Exception:
            continue
        fmt = (cell.number_format or "").lower()

        if '%' in fmt:
            sample = df[col].dropna().head(5).tolist()
            if sample and all(isinstance(v, (int, float)) for v in sample):
                if all(-1 <= v <= 1 for v in sample):
                    traps.append({
                        "type": "percentage_as_fraction",
                        "sheet": sheet_name,
                        "col": col_letter,
                        "name": str(col),
                        "note": "值域 0-1 且 number_format 含 %；公式必须返回小数，不可写 =X*3%",
                        "suggested_read_code": (
                            f'# 列"{col}"存储为 0-1 小数，展示时需 ×100 加 %\n'
                            f'df["{col}_pct"] = df["{col}"] * 100  # 数值运算用原值\n'
                            f'# KPI 展示：f"{{df[\\"{col}\\"].mean() * 100:.1f}}%"'
                        ),
                    })

        if any(tok in fmt for tok in ('yyyy', 'yy', 'mm', 'dd', 'hh')):
            traps.append({
                "type": "date_as_float",
                "sheet": sheet_name,
                "col": col_letter,
                "name": str(col),
                "note": "Excel 日期序列号；pandas 读取用 parse_dates",
                "suggested_read_code": (
                    f'# 列"{col}"是 Excel 日期序列号（类似 45000 代表 2023-03-15）\n'
                    f'df = pd.read_excel("<file>", sheet_name="{sheet_name}", parse_dates=["{col}"])\n'
                    f'# 若已读取但类型不对：df["{col}"] = pd.to_datetime(df["{col}"], unit="D", origin="1899-12-30")'
                ),
            })

    return traps


def _collect_named_ranges(wb) -> list[dict]:
    out: list[dict] = []
    try:
        dn_container = wb.defined_names
        names_iter = dn_container if hasattr(dn_container, '__iter__') else []
        for dn in names_iter:
            try:
                dn_obj = dn_container[dn] if hasattr(dn_container, '__getitem__') else dn
            except Exception:
                dn_obj = dn
            refers = (getattr(dn_obj, 'value', None)
                      or getattr(dn_obj, 'attr_text', None)
                      or str(dn_obj))
            out.append({"name": str(dn), "refers_to": str(refers)})
    except Exception:
        pass
    return out


# --- fuzzy file resolver (救回 Windows cwd 漂移 / 相对路径错 / 空格大小写差异) ---

# 搜索的常见子目录（相对 cwd）。顺序代表 prefer。
_FUZZY_SEARCH_SUBDIRS = ("", "input", "data", "workspace", "files", "uploads", "..")
# 候选最多返回个数（防 hint 过长）
_FUZZY_MAX_CANDIDATES = 5
# 名字相似度阈值（difflib SequenceMatcher.ratio）— 太低会乱匹配，太高漏 "订单 表.xlsx"
_FUZZY_MIN_RATIO = 0.78
# 同后缀文件列举上限（防止超大目录扫穿）
_FUZZY_MAX_LIST_PER_DIR = 200

# 支持后缀（与本 SKILL 处理范围一致）
_FUZZY_EXTS = {".xlsx", ".xlsm", ".csv", ".tsv"}


def _normalize_name(name: str) -> str:
    """规范化文件名用于相似度比较：去后缀、小写、压缩空白。"""
    import re as _re
    stem = Path(name).stem.lower()
    # 把任何连续 whitespace（含全角空格 　）压成单空格再去掉
    stem = _re.sub(r"[\s　]+", "", stem)
    return stem


def _fuzzy_find_candidates(requested: str, cwd: Path | None = None) -> list[dict]:
    """文件找不到时，在 cwd 及常见子目录里 fuzzy 匹配同后缀的近似文件。

    匹配规则（按优先级）：
      1. 绝对路径解析后存在（处理 cwd 漂移：相对路径找不到但绝对路径存在）
      2. basename 完全一致（仅 cwd 不同）
      3. 去空格 + 大小写不敏感一致（"订单 表.xlsx" vs "订单表.xlsx"）
      4. difflib ratio >= _FUZZY_MIN_RATIO 的近似匹配
    返回 [{"path": "...", "score": 0.0-1.0, "reason": "..."}]，按 score 降序。
    """
    import difflib
    from os import getcwd

    cwd = cwd or Path(getcwd())
    req_path = Path(requested)
    req_name = req_path.name
    req_ext = req_path.suffix.lower()
    req_norm = _normalize_name(req_name)

    # 限定只搜支持的后缀；若请求的后缀不在白名单（如无后缀），匹配所有支持后缀
    target_exts = {req_ext} if req_ext in _FUZZY_EXTS else _FUZZY_EXTS

    # 收集搜索目录（去重 + 实际存在的）
    search_dirs: list[Path] = []
    seen: set[Path] = set()
    for sub in _FUZZY_SEARCH_SUBDIRS:
        d = (cwd / sub).resolve() if sub else cwd.resolve()
        if d in seen or not d.is_dir():
            continue
        seen.add(d)
        search_dirs.append(d)

    # 若请求是相对路径，把它的父目录也加进来（处理 "input/x.xlsx" 但 cwd 已是 input/ 的情形）
    if not req_path.is_absolute() and req_path.parent != Path("."):
        # 模型可能写 "input/订单.xlsx"，实际文件就在 cwd 根目录 —— search_dirs 已覆盖 cwd。
        # 也可能写 "订单.xlsx"，但实际在 ./input/ —— 已被 _FUZZY_SEARCH_SUBDIRS 覆盖。
        pass

    candidates: list[dict] = []
    for d in search_dirs:
        try:
            entries = list(d.iterdir())[:_FUZZY_MAX_LIST_PER_DIR]
        except (OSError, PermissionError):
            continue
        for entry in entries:
            if not entry.is_file():
                continue
            if entry.suffix.lower() not in target_exts:
                continue
            # 完全一致（basename 匹配，只是位置不同）
            if entry.name == req_name:
                candidates.append({
                    "path": str(entry.resolve().as_posix()),
                    "score": 1.0,
                    "reason": "basename 完全一致（cwd 漂移）",
                })
                continue
            # 大小写 / 空格不敏感一致
            entry_norm = _normalize_name(entry.name)
            if entry_norm == req_norm:
                candidates.append({
                    "path": str(entry.resolve().as_posix()),
                    "score": 0.95,
                    "reason": "去空格/大小写后一致",
                })
                continue
            # difflib 相似度
            ratio = difflib.SequenceMatcher(None, entry_norm, req_norm).ratio()
            if ratio >= _FUZZY_MIN_RATIO:
                candidates.append({
                    "path": str(entry.resolve().as_posix()),
                    "score": round(ratio, 3),
                    "reason": f"名称相似度 {ratio:.0%}",
                })

    # 去重（不同搜索目录可能解析到同一文件） + 按 score 降序
    by_path: dict[str, dict] = {}
    for c in candidates:
        prev = by_path.get(c["path"])
        if prev is None or c["score"] > prev["score"]:
            by_path[c["path"]] = c
    out = sorted(by_path.values(), key=lambda x: -x["score"])
    return out[:_FUZZY_MAX_CANDIDATES]


# --- 加密 / OLE 容器检测（防 case_031/058 类反复装解密库死循环）---
# 历史评测里加密 xlsx 单 trial 平均 22 tool call、最高 46 — 模型反复 pip install
# msoffcrypto-tool/pyzipper 试图破解。直接在 inspect 入口拦截 → stop_and_report。

# Magic 字节：
#   OLE Compound File Binary (CFB) — 老式加密 xlsx / .xls 共用
_OLE_CFB_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
#   ZIP 容器 — OOXML 加密包从这里继续探测 namelist 是否有 EncryptedPackage
_ZIP_MAGIC = b'PK\x03\x04'


def _detect_encrypted_or_ole(filepath: Path) -> str | None:
    """识别加密 xlsx / OLE 容器文件。

    返回:
      - 'ole_cfb'         : OLE Compound 容器（加密 xlsx 或老 .xls 等）
      - 'encrypted_ooxml' : OOXML 加密包（ZIP 容器但里面只有 EncryptedPackage）
      - None              : 普通文件，让后续流程处理

    不在此函数处理 BadZipFile / 普通 xlsx — 命中加密时返回标签即可，
    miss 时让 inspect 后续路径继续走（包括 BadZipFile 兜底）。
    """
    try:
        with open(filepath, "rb") as f:
            head = f.read(8)
    except OSError:
        return None
    if head == _OLE_CFB_MAGIC:
        return "ole_cfb"
    if head[:4] == _ZIP_MAGIC:
        try:
            with zipfile.ZipFile(filepath, "r") as zf:
                names = set(zf.namelist())
            # OOXML 加密包标识 — EncryptedPackage + EncryptionInfo 至少一个存在
            # 且 xl/workbook.xml 不存在（普通 xlsx 必然有 xl/workbook.xml）
            if ("EncryptedPackage" in names or "EncryptionInfo" in names) \
                    and not any(n.startswith("xl/workbook") for n in names):
                return "encrypted_ooxml"
        except (zipfile.BadZipFile, OSError):
            return None
    return None


def inspect_file(filepath: str) -> dict:
    """v4 INSPECT: produce full JSON report."""
    path = Path(filepath)

    if not path.exists():
        candidates = _fuzzy_find_candidates(filepath)
        result = {
            "ok": False,
            "error_class": "file_not_found",
            "human_summary": f"文件不存在: {filepath}",
            "suggested_fix": {"action": "ask_user", "reason": "请确认文件路径"},
        }
        if candidates:
            top = candidates[0]
            result["fuzzy_candidates"] = candidates
            result["human_summary"] += (
                f"。在附近找到 {len(candidates)} 个近似文件，"
                f"最像的是: {top['path']}（{top['reason']}）"
            )
            result["suggested_fix"] = {
                "action": "retry_with_path",
                "reason": "原路径找不到，但在 cwd 附近匹配到近似文件，建议改用绝对路径重试",
                "params": {
                    "suggested_path": top["path"],
                    "all_candidates": candidates,
                },
            }
        return result

    abs_path = str(path.resolve().as_posix())
    ext = path.suffix.lower()
    file_size = path.stat().st_size

    # Step 0 路由白名单：本技能只支持 .xlsx / .xlsm / .csv / .tsv。
    # 其他格式一律拒绝并 stop_and_report——不要在本流程内强行解析或替代分析
    # （实测 case_028/041/042/059 都是字面遵守"提示"后仍"帮忙"处理）。
    # 措辞中性：调用方可能是模型/Agent 链路，不一定是真人；由调用方决定后续。
    UNSUPPORTED_FORMATS = {
        ".xls":  "Excel 97-2003 旧格式，需先转换为 .xlsx",
        ".xlsb": "Excel 二进制格式，需先转换为 .xlsx",
        ".db":   "SQLite 数据库不在支持范围（本技能仅处理表格文件），需先导出为 .xlsx/.csv",
        ".sqlite": "SQLite 数据库不在支持范围，需先导出为 .xlsx/.csv",
        ".sqlite3": "SQLite 数据库不在支持范围，需先导出为 .xlsx/.csv",
        ".zip":  "压缩包不在支持范围，需先解压后重新提供内部的 .xlsx/.csv 文件",
        ".pdf":  "PDF 不在支持范围，需先转换为表格（.xlsx/.csv）",
        ".docx": "Word 文档不在支持范围，需先把数据导出为 .xlsx/.csv",
        ".doc":  "Word 文档不在支持范围，需先把数据导出为 .xlsx/.csv",
        ".pptx": "PPT 不在支持范围",
        ".txt":  "纯文本不在支持范围，如果是分隔符表格请改后缀为 .csv 或 .tsv",
    }
    if ext in UNSUPPORTED_FORMATS:
        return {
            "ok": False,
            "error_class": "unsupported_format",
            "human_summary": f"{ext} 格式不支持：{UNSUPPORTED_FORMATS[ext]}",
            "suggested_fix": {
                "action": "stop_and_report",
                "reason": "本技能只处理 .xlsx / .xlsm / .csv / .tsv",
                "params": {
                    "required_input_change": UNSUPPORTED_FORMATS[ext],
                    "must_not": [
                        "不要 pip install xlrd / pypdf / pdfplumber / sqlite3 等任何依赖来强行解析",
                        "不要写转换脚本（convert_xls / extract_pdf / unzip 等）",
                        "不要拿其他文件作为替代分析",
                        "不要执行后续 CLASSIFY/ANALYZE/REPORT 任何步骤",
                    ],
                },
            },
        }

    # 加密 / 密码保护文件兜底（防 case_031/058 类反复装 msoffcrypto/pyzipper）
    # 走在 UNSUPPORTED_FORMATS 之后：.xls 等老格式已被前面拦掉；这里只剩 .xlsx/.xlsm
    # 但容器实际是加密的情况
    encryption_kind = _detect_encrypted_or_ole(path)
    if encryption_kind:
        kind_label = {
            "ole_cfb": "OLE 加密容器（CFB 格式，常见于设了打开密码的 xlsx）",
            "encrypted_ooxml": "OOXML 加密包（EncryptedPackage 容器）",
        }[encryption_kind]
        return {
            "ok": False,
            "error_class": "encrypted_or_password_protected",
            "human_summary": f"文件被加密：{kind_label}，无法直接读取",
            "suggested_fix": {
                "action": "stop_and_report",
                "reason": "本技能不破解加密文件，需用户提供解密版本",
                "params": {
                    "required_input_change": (
                        "请用 Excel 打开该文件，输入密码后选『另存为 → 工具 → 常规选项 → "
                        "把『打开权限密码』清空 → 保存』生成解密版本再重新提供。"
                    ),
                    "must_not": [
                        "不要 pip install msoffcrypto-tool / pyzipper / defusedxml 等任何解密依赖",
                        "不要尝试暴力破解密码或猜测密码",
                        "不要拿其他文件作为替代分析",
                        "不要执行后续 CLASSIFY/ANALYZE/REPORT 任何步骤",
                    ],
                },
            },
        }

    file_size_mb = round(file_size / (1024 * 1024), 1)
    result: dict = {
        "ok": True,
        "file": abs_path,
        "file_size": file_size,
        "file_size_mb": file_size_mb,
        "format": ext.lstrip('.'),
        "human_summary": "",
        "sheet_names": [],
        "sheet_dims": {},
        "column_info": [],
        "named_ranges": [],
        "formula_layer": {"count": 0, "functions_used": [], "cross_sheet_refs": 0},
        "volatility": {"has_vba": False, "has_external_links": False, "has_pivot": False},
        "data_traps": [],
        "previews": {},
        "profile_scope": {},
        "blocking_issues": [],
        "next_action": {"code": "proceed", "hint": ""},
    }

    # 大文件预警：> LARGE_FILE_WARN_MB 时加 trap，提示后续 analyze 不要直接全量
    if file_size_mb > LARGE_FILE_WARN_MB:
        result["data_traps"].append({
            "type": "large_file",
            "file_size_mb": file_size_mb,
            "note": (
                f"文件 {file_size_mb:.1f} MB（> {LARGE_FILE_WARN_MB} MB）。"
                f"后续 analyze 步骤直接全量 read 可能耗时长/吃内存；"
                f"建议显式 `xlsx_analyze.py --head N` 或 `--profile-only`，"
                f"或在 Step 2 用 chunked pandas。"
                f"xlsx_analyze 在 > 200 MB 时会自动 fail-fast。"
            ),
            "suggested_read_code": (
                "# 大文件读取选一：\n"
                "# A) 仅画像（最快）：\n"
                f"#    python xlsx_analyze.py '{path.name}' --profile-only\n"
                "# B) 读前 N 行做样本分析：\n"
                f"#    python xlsx_analyze.py '{path.name}' --head 100000\n"
                "# C) 自行 chunked 聚合：\n"
                "import pandas as pd\n"
                "agg = {}\n"
                f"for chunk in pd.read_csv('{path.name}', chunksize=200_000):\n"
                "    # 在此累加你的 groupby 结果\n"
                "    pass\n"
            ),
        })

    # --- CSV / TSV ---
    if ext in ('.csv', '.tsv'):
        sep = '\t' if ext == '.tsv' else ','
        enc = _sniff_csv_encoding(filepath)
        if enc is None:
            return {
                "ok": False,
                "error_class": "csv_encoding_error",
                "human_summary": "CSV 编码识别失败（试过 utf-8-sig/utf-8/gbk/gb18030）",
                "suggested_fix": {
                    "action": "ask_user",
                    "reason": "编码非常见 Windows 中文编码",
                    "params": {"--encoding": "<需要用户指定>"},
                },
            }
        result["encoding"] = enc
        # 行数统计：小文件精确数；> CSV_EXACT_COUNT_LIMIT_MB 改用字节估算（避免 30s+ IO）
        file_size_mb = file_size / (1024 * 1024)
        row_count_estimated = False
        if file_size_mb > CSV_EXACT_COUNT_LIMIT_MB:
            real_row_count, row_count_estimated = _estimate_csv_rows(
                filepath, enc, file_size)
        else:
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    real_row_count = sum(1 for _ in f) - 1
                real_row_count = max(real_row_count, 0)
            except Exception:
                real_row_count = -1
        try:
            df = pd.read_csv(filepath, sep=sep, encoding=enc,
                             nrows=INSPECT_SAMPLE_ROWS)
        except Exception as e:
            return {
                "ok": False,
                "error_class": "csv_encoding_error",
                "human_summary": f"CSV 读取失败: {e}",
                "suggested_fix": {"action": "ask_user", "reason": str(e)},
            }
        sheet_name = path.stem
        result["sheet_names"] = [sheet_name]
        row_count = real_row_count if real_row_count >= 0 else int(len(df))
        result["sheet_dims"] = {sheet_name: [row_count, int(len(df.columns))]}
        if row_count_estimated:
            result["row_count_estimated"] = True
        for col_idx, col in enumerate(df.columns, start=1):
            s = df[col].dropna()
            # name_repr 不在此重复：高危列名(隐藏字符/前后空白/中英间空格等)的 repr
            # 由 _detect_column_name_quirks 产出的 column_name_quirk trap 携带(含 repr +
            # 修复代码，且只对高危列触发)，避免与那套检测重复/漂移。
            result["column_info"].append({
                "sheet": sheet_name,
                "col": _col_letter(col_idx),
                "name": str(col),
                "dtype": str(df[col].dtype),
                "non_null": int(s.count()),
                "samples": [str(v) for v in s.head(3).tolist()],
            })
        rows_profiled = int(len(df))
        result["profile_scope"][sheet_name] = _profile_scope_entry(
            row_count, rows_profiled)
        sampled = result["profile_scope"][sheet_name]["sampled"]
        if sampled:
            result["data_traps"].append(
                _inspect_sampled_trap(sheet_name, row_count, rows_profiled))
        # 取 raw 前 10 行 (header=None) 给 _detect_skiprows_from_rows 用
        _raw_head_rows = None
        try:
            _raw_df = pd.read_csv(
                filepath, sep=sep, encoding=enc, header=None, nrows=10,
                dtype=str, keep_default_na=False,
            )
            _raw_head_rows = _raw_df.values.tolist()
        except Exception:
            _raw_head_rows = None
        result["data_traps"].extend(_detect_column_name_quirks(sheet_name, df, raw_head_rows=_raw_head_rows))
        if not df.empty:
            # 真实表尾：被截断且行数精确时用 skiprows 读尾部，否则用 df 尾部
            csv_truncated = (sampled and real_row_count >= 0
                             and not row_count_estimated)
            tail = []
            if csv_truncated:
                try:
                    _tail_df = pd.read_csv(
                        filepath, sep=sep, encoding=enc, header=None,
                        names=list(df.columns),
                        skiprows=range(1, max(1, real_row_count - 5 + 1)),
                    )
                    tail = _tail_df.tail(5).astype(str).values.tolist()
                except Exception:
                    tail = []
            if not tail:
                tail = (df.tail(5).astype(str).values.tolist()
                        if len(df) > 5 else [])
            result["previews"][sheet_name] = {
                "head": df.head(5).astype(str).values.tolist(),
                "tail": tail,
            }
        sample_note = (f"（列画像采样前 {rows_profiled} 行）" if sampled else "")
        result["human_summary"] = (
            f"CSV，编码 {enc}，{row_count} 行 × {len(df.columns)} 列{sample_note}"
        )
        result["next_action"]["hint"] = "进入 Step 2 CLASSIFY，判断任务归属 7 类"
        result["next_action"]["script_safety"] = _SCRIPT_SAFETY_HINT
        validate_hint = _recommend_validate_columns(result["column_info"])
        if validate_hint:
            result["next_action"]["recommend_validate_columns"] = validate_hint
            result["next_action"]["hint"] += "\n" + validate_hint["hint"]
        return result

    # --- XLSX / XLSM ---
    if ext not in ('.xlsx', '.xlsm'):
        return {
            "ok": False,
            "error_class": "unsupported_format",
            "human_summary": f"不支持的格式: {ext}",
            "suggested_fix": {
                "action": "stop_and_report",
                "reason": "仅支持 .xlsx/.xlsm/.csv/.tsv",
            },
        }

    has_vba, has_ext, has_pivot = _check_vba_external(filepath)
    result["volatility"] = {
        "has_vba": has_vba,
        "has_external_links": has_ext,
        "has_pivot": has_pivot,
    }
    if has_vba:
        result["data_traps"].append({
            "type": "has_vba",
            "note": ".xlsm 含 VBA；编辑必须走 xlsx_unpack→XML→xlsx_pack，禁止 openpyxl.save() 否则丢 VBA",
            "suggested_read_code": "# 读数据仍可正常：\n# df = pd.read_excel('<file>', sheet_name='<sheet>')\n# 仅在需要编辑/另存时才有风险，分析场景可忽略",
        })
    if has_ext:
        result["data_traps"].append({
            "type": "external_link",
            "note": "工作簿含外部链接引用；如实告知用户，不自动解析",
            "suggested_read_code": "# 外部链接列读到的是缓存值（可能过期）。\n# 若值为 None 或 #REF!，向用户说明无法解析，不要猜测",
        })

    # 大文件分流：> STREAMING_INSPECT_MB 走 read_only=True 流式（省内存/省时，避免被
    # 1 分钟前台预算背景化）；小文件保持 read_only=False 全功能（零回归、原生合并单元格）。
    streaming = file_size_mb > STREAMING_INSPECT_MB
    scan_deadline = time.monotonic() + INSPECT_SCAN_BUDGET_SEC
    # 大文件起心跳：load_workbook 及后续扫描中途不打印，靠它每 20s 向 stderr 打一行，
    # 重置 Shell 命令的 60s idle 计时器，避免被切后台（见 _Heartbeat 注释）。
    _hb = (_Heartbeat(f'{path.name} {file_size_mb:.0f}MB').start()
           if file_size_mb > HEARTBEAT_MIN_MB else None)
    try:
        wb = load_workbook(filepath, data_only=False, read_only=streaming)
    except PermissionError:
        if _hb: _hb.stop()
        return {
            "ok": False,
            "error_class": "file_locked_by_excel",
            "human_summary": "文件被占用（可能 Excel 正在打开）",
            "suggested_fix": {
                "action": "ask_user",
                "reason": "Windows 下 Excel 打开会锁文件",
                "params": {"hint": "请先在 Excel 关闭该文件再重试"},
            },
        }
    except zipfile.BadZipFile:
        if _hb: _hb.stop()
        return {
            "ok": False,
            "error_class": "zip_corrupted",
            "human_summary": "文件 ZIP 结构损坏",
            "suggested_fix": {
                "action": "stop_and_report",
                "reason": "进入 §L2.REPAIR 剧本，用 xlsx_unpack.py 诊断",
            },
        }
    except Exception as e:
        if _hb: _hb.stop()
        return {
            "ok": False,
            "error_class": "xml_malformed",
            "human_summary": f"工作簿加载失败: {e}",
            "suggested_fix": {"action": "stop_and_report", "reason": str(e)},
        }

    result["sheet_names"] = list(wb.sheetnames)

    result["named_ranges"] = _collect_named_ranges(wb)
    if result["named_ranges"]:
        result["data_traps"].append({
            "type": "named_range_dependency",
            "count": len(result["named_ranges"]),
            "note": "存在命名区域；删列/改名会级联坏，EDIT 前必须核对",
            "suggested_read_code": "# 分析场景下命名区域不影响读取；仅在 EDIT 时才需警惕",
        })

    # 大文件流式模式：诚实告知未做合并单元格/多级表头探测，并给按需补检命令
    if streaming:
        result["streaming_inspect"] = True
        result["data_traps"].append({
            "type": "large_file_streaming",
            "file_size_mb": file_size_mb,
            "note": (
                f"文件 {file_size_mb:.1f} MB（> {STREAMING_INSPECT_MB} MB），INSPECT 走"
                f"流式低内存模式。**未做合并单元格 / 多级表头探测**（流式下不可用），"
                f"行数可能未精确统计。列画像/预览取自前 {INSPECT_SAMPLE_ROWS} 行采样。"
                f"百分比/日期/公式错等其余 trap 正常。"
            ),
            "skipped": ["merged_cell", "multi_level_header"],
            "suggested_read_code": (
                "# 若怀疑该表有合并单元格 / 多级表头（典型：报表型 Excel），按需单独检测：\n"
                f"#   python xlsx_reader.py '{Path(filepath).name}' --merged-scan\n"
                "# 若确认是多级表头，再用 header=[0,1] 读：\n"
                f"#   df = pd.read_excel('<file>', sheet_name='<sheet>', header=[0,1])\n"
                "# 否则按普通单层表头处理即可。"
            ),
        })

    result["formula_layer"] = _extract_formula_layer(wb, deadline=scan_deadline)
    if result["formula_layer"]["cross_sheet_refs"] > 0:
        result["data_traps"].append({
            "type": "cross_sheet_ref",
            "count": result["formula_layer"]["cross_sheet_refs"],
            "note": "跨 sheet 引用；汇总 sheet 公式可能未缓存（cache 为 None）。先尝试 data_only=True 读 cache；若 cache 全空，立即降级到从明细 sheet pandas 重算，不要在自动化场景里建议用户另存",
            "suggested_read_code": (
                "# Step 1: 尝试 openpyxl data_only=True 读公式缓存值\n"
                "import openpyxl, pandas as pd\n"
                "wb = openpyxl.load_workbook('<file>', data_only=True)\n"
                "ws = wb['<汇总 sheet 名>']\n"
                "data = [[c.value for c in row] for row in ws.iter_rows()]\n"
                "df = pd.DataFrame(data[1:], columns=data[0])\n"
                "\n"
                "# Step 2: 若 df 数值列大量为 None/NaN（源文件从未被 Excel 打开/保存过，cache 未生成）\n"
                "# → 自动化场景（评测 / CI / 服务端）下用户无法重新另存，立即放弃 cache 读法，\n"
                "#   改用 pandas 从明细 sheet 重新聚合，按汇总 sheet 列含义匹配 group_key / value 列：\n"
                "detail = pd.read_excel('<file>', sheet_name='<明细 sheet 名>')\n"
                "# summary = detail.groupby('<group_key_col>')['<value_col>'].sum().reset_index()\n"
                "# 仅在交互式使用且用户能操作时，才提示『请用户在 Excel 中打开后另存一次』作为可选路径"
            ),
        })

    da_hits = sorted(set(result["formula_layer"]["functions_used"]) & DYNAMIC_ARRAY_FUNCS)
    if da_hits:
        result["data_traps"].append({
            "type": "dynamic_array_func",
            "functions": da_hits,
            "note": "recalc.py 对动态数组函数返回 status=partial（非 fail），可继续 deliver",
            "suggested_read_code": (
                "# 动态数组函数（SORT/FILTER/UNIQUE/XLOOKUP 等）在 openpyxl data_only=True 下\n"
                "# 可能只返回左上角值。若列长度异常短，改用 openpyxl data_only=False + 手动求值\n"
                "# 或请用户先在 Excel 中打开另存以 spill 缓存"
            ),
        })

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_count_estimated = False
        try:
            if streaming:
                # 单次解析：从同一个 read_only wb 的 iter_rows 取样，不再二次 pd.read_excel
                df, _ = _sample_df_via_openpyxl(ws, INSPECT_SAMPLE_ROWS)
            else:
                df = pd.read_excel(filepath, sheet_name=sheet_name,
                                   nrows=INSPECT_SAMPLE_ROWS)
        except Exception as e:
            df = pd.DataFrame()
            result["data_traps"].append({
                "type": "unreadable_sheet",
                "sheet": sheet_name,
                "note": f"Sheet '{sheet_name}' 读取失败: {type(e).__name__}: {e}",
                "suggested_read_code": f"# 跳过此 sheet 或尝试 openpyxl 直接读取\n"
                                       f"# ws = openpyxl.load_workbook('{path.name}')['{sheet_name}']\n"
                                       f"# data = [[c.value for c in row] for row in ws.iter_rows()]",
            })

        # Use openpyxl max_row for real row count (not truncated nrows=2000)
        # 流式(read_only)下 max_row 可能为 None（无 dimension 标签）→ 退化为采样行数并标记 estimated
        try:
            _mr = ws.max_row
        except Exception:
            _mr = None
        if _mr is None:
            real_rows = int(len(df))
            if streaming:
                row_count_estimated = True
        else:
            real_rows = int(_mr)
            if real_rows > 0:
                real_rows -= 1  # subtract header row
            real_rows = max(real_rows, int(len(df)))
        dims = [real_rows, int(len(df.columns))]
        result["sheet_dims"][sheet_name] = dims
        if row_count_estimated:
            result.setdefault("row_count_estimated_sheets", []).append(sheet_name)

        for col_idx, col in enumerate(df.columns, start=1):
            s = df[col].dropna()
            # name_repr 不在此重复：高危列名(隐藏字符/前后空白/中英间空格等)的 repr
            # 由 _detect_column_name_quirks 产出的 column_name_quirk trap 携带(含 repr +
            # 修复代码，且只对高危列触发)，避免与那套检测重复/漂移。
            result["column_info"].append({
                "sheet": sheet_name,
                "col": _col_letter(col_idx),
                "name": str(col),
                "dtype": str(df[col].dtype),
                "non_null": int(s.count()),
                "samples": [str(v) for v in s.head(3).tolist()],
            })

        traps = _detect_sheet_traps(sheet_name, df, ws, streaming=streaming)
        result["data_traps"].extend(traps)
        result["data_traps"].extend(_detect_duplicate_headers(sheet_name, ws))
        # 取 raw 前 10 行 (header=None 等价: 直接读 openpyxl cell 值) 给 _detect_skiprows 用
        _raw_head_rows = None
        try:
            _raw_head_rows = [
                list(row) for row in ws.iter_rows(
                    min_row=1, max_row=10, values_only=True)
            ]
        except Exception:
            _raw_head_rows = None
        result["data_traps"].extend(
            _detect_column_name_quirks(sheet_name, df, raw_head_rows=_raw_head_rows))
        result["data_traps"].extend(
            _detect_formula_errors(sheet_name, ws, deadline=scan_deadline))

        if not df.empty:
            rows_profiled = int(len(df))
            result["profile_scope"][sheet_name] = _profile_scope_entry(
                real_rows, rows_profiled)
            if result["profile_scope"][sheet_name]["sampled"]:
                # 被采样：补 inspect_sampled trap，且 tail 取真实表尾（非采样段尾巴）
                result["data_traps"].append(
                    _inspect_sampled_trap(sheet_name, real_rows, rows_profiled))
                # 流式(read_only)下取真实表尾要整表重扫一遍（慢 + max_row 可能 None）→
                # 退化为采样段尾部，避免顶满 1 分钟前台预算。
                tail = ([] if streaming
                        else _real_tail_xlsx(ws, len(df.columns), 5))
            else:
                tail = (df.tail(5).astype(str).values.tolist()
                        if len(df) > 5 else [])
            result["previews"][sheet_name] = {
                "head": df.head(5).astype(str).values.tolist(),
                "tail": tail,
            }

    # 跨 sheet 同名列 → sheet_column_ambiguity
    # 场景：case_061 销售额_含税 / 销售额_不含税 两个 sheet 都有"销售额"列，
    # 若 KPI 不标口径会混淆。从 sheet 名提取差异化后缀供模型标题使用。
    if len(wb.sheetnames) >= 2:
        from collections import defaultdict
        col_to_sheets = defaultdict(list)
        for ci in result["column_info"]:
            col_to_sheets[ci["name"]].append(ci["sheet"])
        ambiguous = {col: sheets for col, sheets in col_to_sheets.items()
                     if len(sheets) >= 2 and len(set(sheets)) >= 2}
        if ambiguous:
            # 推导差异化后缀：各 sheet 名按 _/-/空格 分词，从尾部找不与其它 sheet
            # 精确重合的 part（如 "销售额_含税" vs "销售额_不含税" → "含税" vs "不含税"）
            def _derive_suffix(sheet_name, others):
                parts = re.split(r'[_\-\s]+', sheet_name)
                other_parts_sets = [set(re.split(r'[_\-\s]+', o)) for o in others]
                for p in reversed(parts):
                    if p and not any(p in ops for ops in other_parts_sets):
                        return p
                return sheet_name  # fallback: 整个 sheet 名
            mappings = {}
            for col, sheets in ambiguous.items():
                mappings[col] = {
                    s: _derive_suffix(s, [o for o in sheets if o != s])
                    for s in sheets
                }
            result["data_traps"].append({
                "type": "sheet_column_ambiguity",
                "ambiguous_columns": list(ambiguous.keys()),
                "sheet_suffix_map": mappings,
                "note": (
                    f"检测到 {len(ambiguous)} 个列在多个 sheet 中同名（"
                    + ", ".join(f"「{c}」" for c in ambiguous.keys())
                    + "），口径可能不同；KPI 标签和图表标题必须带 sheet 口径后缀以避免混淆。"
                    + "后缀映射见 sheet_suffix_map"
                ),
                "suggested_read_code": (
                    "# 多 sheet 同名列 → 读取时用 sheet_name=None 一次读全部，再加后缀区分：\n"
                    "import pandas as pd\n"
                    "sheets = pd.read_excel('<file>', sheet_name=None)\n"
                    "# 在 KPI / 图表标题中显式标注来源 sheet：\n"
                    "# 例如 ready_kpi.label = '销售额（含税）' 而非 '销售额'"
                ),
            })

    total_rows = sum(d[0] for d in result["sheet_dims"].values())
    trap_count = len(result["data_traps"])
    high_risk = []
    if has_vba:
        high_risk.append("含 VBA")
    if has_ext:
        high_risk.append("含外链")
    if has_pivot:
        high_risk.append("含 pivot")
    risk_note = f"（高风险: {', '.join(high_risk)}）" if high_risk else ""
    sampled_sheets = [s for s, v in result["profile_scope"].items()
                      if v.get("sampled")]
    sample_note = (
        f"（{len(sampled_sheets)} 个 sheet 列画像采样前 {INSPECT_SAMPLE_ROWS} 行）"
        if sampled_sheets else ""
    )
    result["human_summary"] = (
        f"{len(result['sheet_names'])} sheet，{total_rows} 行，"
        f"{result['formula_layer']['count']} 公式，{trap_count} 个 data trap"
        f"{risk_note}{sample_note}"
    )
    result["next_action"]["hint"] = (
        "进入 Step 2 CLASSIFY；若 data_traps 非空，后续处理必须按每条 trap 的 note 执行"
    )
    result["next_action"]["script_safety"] = _SCRIPT_SAFETY_HINT
    validate_hint = _recommend_validate_columns(result["column_info"])
    if validate_hint:
        result["next_action"]["recommend_validate_columns"] = validate_hint
        result["next_action"]["hint"] += "\n" + validate_hint["hint"]
    # 流式(read_only) wb 必须 close，否则文件句柄不释放（Windows 会锁文件）
    if streaming:
        try:
            wb.close()
        except Exception:
            pass
    if _hb: _hb.stop()
    return result


def merged_scan(filepath: str) -> dict:
    """按需检测合并单元格 / 多级表头（供大文件流式 inspect 后的补检入口）。
    用 iterparse 流式扫 sheet XML，低内存；大文件 inspect 默认不跑这步（耗时），
    模型按 large_file_streaming trap 的提示在怀疑多级表头时显式调用。"""
    p = Path(filepath)
    if not p.exists():
        return {"ok": False, "error_class": "file_not_found",
                "human_summary": f"文件不存在: {filepath}"}
    if p.suffix.lower() not in ('.xlsx', '.xlsm'):
        return {"ok": False, "error_class": "unsupported_format",
                "human_summary": f"--merged-scan 仅支持 .xlsx/.xlsm，当前 {p.suffix}"}
    try:
        per_xml = _iterparse_merged_cells(filepath)
    except Exception as e:
        return {"ok": False, "error_class": "merged_scan_failed",
                "human_summary": f"合并单元格扫描失败: {type(e).__name__}: {e}"}
    # sheet XML → workbook sheet 名（best-effort，按顺序映射）
    try:
        wb = load_workbook(filepath, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
    except Exception:
        names = []
    by_sheet = {}
    for i, (_xml, refs) in enumerate(per_xml.items()):
        label = names[i] if i < len(names) else _xml
        if refs:
            by_sheet[label] = refs
    total = sum(len(v) for v in per_xml.values())
    has_merged = total > 0
    return {
        "ok": True,
        "file": str(p.resolve()),
        "has_merged_cells": has_merged,
        "total_merged_ranges": total,
        "by_sheet": by_sheet,
        "human_summary": (
            f"检出 {total} 个合并单元格区域（疑似多级表头/报表结构）"
            if has_merged else "未检出合并单元格，按普通单层表头处理即可"),
        "next_action": {"code": "proceed", "hint": (
            "若像多级表头: df = pd.read_excel(file, sheet_name='<sheet>', header=[0,1])；"
            "若是数据区合并(分类首格合并): 读后 df = df.ffill()"
            if has_merged else "无合并单元格，正常单层读取即可")},
    }


# =========================================================================
# Legacy modes (preserved for backwards compat)
# =========================================================================

def read_file(filepath: str, sheet: str = None):
    path = Path(filepath)
    ext = path.suffix.lower()
    if ext in ('.csv', '.tsv'):
        sep = '\t' if ext == '.tsv' else ','
        encoding = _sniff_csv_encoding(filepath) or 'utf-8'
        df = pd.read_csv(filepath, sep=sep, encoding=encoding)
        return {path.stem: df}
    elif ext in ('.xlsx', '.xlsm', '.xls'):
        if sheet:
            df = pd.read_excel(filepath, sheet_name=sheet)
            return {sheet: df}
        return pd.read_excel(filepath, sheet_name=None)
    raise ValueError(f"Unsupported file format: {ext}")


def quality_audit(df) -> dict:
    report = {
        'shape': {'rows': len(df), 'columns': len(df.columns)},
        'columns': list(df.columns),
        'dtypes': df.dtypes.astype(str).to_dict(),
        'missing': {col: int(n) for col, n in df.isnull().sum().items() if n > 0},
        'missing_pct': {col: round(n / len(df) * 100, 1) for col, n in df.isnull().sum().items() if n > 0},
        'duplicate_rows': int(df.duplicated().sum()),
        'empty_columns': [col for col in df.columns if df[col].isnull().all()],
        'mixed_types': {},
        'potential_issues': [],
    }
    for col in df.select_dtypes(include='object').columns:
        non_null = df[col].dropna()
        if len(non_null) == 0:
            continue
        types_found = non_null.apply(type).unique()
        if len(types_found) > 1:
            report['mixed_types'][col] = [t.__name__ for t in types_found]
    for col in df.select_dtypes(include='number').columns:
        vals = df[col].dropna()
        if len(vals) > 0 and (vals.between(1900, 2100)).all() and vals.nunique() < 10:
            report['potential_issues'].append(f"Column '{col}' may contain years formatted as numbers")
    for col in df.select_dtypes(include='object').columns:
        sample = df[col].dropna().head(100)
        if sample.apply(lambda x: isinstance(x, str) and x != x.strip()).any():
            report['potential_issues'].append(f"Column '{col}' has values with leading/trailing spaces")
    return report


def compute_stats(df) -> dict:
    stats = {}
    for col in df.select_dtypes(include='number').columns:
        s = df[col].dropna()
        if len(s) == 0:
            continue
        stats[col] = {
            'count': int(s.count()),
            'sum': round(float(s.sum()), 4),
            'mean': round(float(s.mean()), 4),
            'median': round(float(s.median()), 4),
            'std': round(float(s.std()), 4) if len(s) > 1 else 0,
            'min': round(float(s.min()), 4),
            'max': round(float(s.max()), 4),
            'nulls': int(df[col].isnull().sum()),
        }
    return stats


def format_preview(df, n: int = 5) -> str:
    with pd.option_context('display.max_columns', None,
                           'display.width', 120,
                           'display.max_colwidth', 30):
        return df.head(n).to_string()


# --- column-name validator: 模型脚本预校验，挡掉中→英幻觉/凭空捏造列 ---

# 模型写完脚本后调用：python xlsx_reader.py file.xlsx --validate-columns 销售额 下单_date
# 对每个用户给的列名：
#   ✓ 存在 → 直接匹配
#   ✗ 不存在 → fuzzy 找最近似列（多策略：原文 + 中文stem + 归一化）
# 不存在且无近似 → 标 "凭空捏造"，建议重看 column_info
_COL_VALIDATE_MIN_RATIO = 0.6

# 模型常自动添加的英文语义后缀 — 用于识别"中→英别名"幻觉
_ALIAS_SUFFIXES = (
    "_date", "_time", "_datetime", "_id", "_no", "_num", "_count",
    "_amount", "_value", "_total", "_sum", "_avg", "_mean",
    "_rate", "_ratio", "_percentage", "_percent", "_pct",
)


def _stem_for_match(name: str) -> tuple[str, str]:
    """返回 (normalized, cjk_only)：
       - normalized: 小写 + 去下划线/空格 + 去常见英文后缀
       - cjk_only: 仅保留中文字符（U+4E00-U+9FFF）
    用于多策略相似度匹配，识别"下单_date" vs "下单日期"这类别名幻觉。
    """
    s = str(name).lower().strip()
    for suf in _ALIAS_SUFFIXES:
        if s.endswith(suf):
            s = s[:-len(suf)]
            break
    norm = re.sub(r"[_\s\-]+", "", s)
    cjk = "".join(c for c in name if "一" <= c <= "鿿")
    return norm, cjk


def _col_similarity(a: str, b: str) -> tuple[float, str]:
    """多策略列名相似度：
       1. 原文 SequenceMatcher.ratio
       2. 归一化（去英文后缀+下划线）后 ratio
       3. 中文 stem substring 互含（'下单' ⊂ '下单日期' 直接给 0.9）
       取最大值 + 标注命中策略。
    """
    import difflib
    raw = difflib.SequenceMatcher(None, a, b).ratio()
    a_norm, a_cjk = _stem_for_match(a)
    b_norm, b_cjk = _stem_for_match(b)
    norm = difflib.SequenceMatcher(None, a_norm, b_norm).ratio() if a_norm and b_norm else 0.0
    cjk = 0.0
    if a_cjk and b_cjk and (a_cjk in b_cjk or b_cjk in a_cjk):
        # 中文 stem 一方包含另一方，给高分
        shorter = min(len(a_cjk), len(b_cjk))
        longer = max(len(a_cjk), len(b_cjk))
        cjk = 0.7 + 0.3 * (shorter / longer)
    scores = [("raw", raw), ("normalized", norm), ("cjk_substring", cjk)]
    best = max(scores, key=lambda x: x[1])
    return best[1], best[0]


def validate_columns(filepath: str, requested: list[str]) -> dict:
    """校验请求的列名是否存在于文件 schema 中；不存在则给最近似建议。"""

    path = Path(filepath)
    if not path.exists():
        # 顺手做文件 fuzzy（与 inspect 一致体验）
        cands = _fuzzy_find_candidates(filepath)
        return {
            "ok": False,
            "error_class": "file_not_found",
            "human_summary": f"文件不存在: {filepath}",
            "fuzzy_file_candidates": cands,
        }

    try:
        sheets = read_file(str(path), sheet=None)
    except Exception as e:
        return {
            "ok": False,
            "error_class": "read_failed",
            "human_summary": f"读取失败：{e}",
        }

    # 汇总所有 sheet 的列名（带 sheet 标注，便于 disambiguate）
    schema = {}  # name → list of sheet names
    for sn, df in sheets.items():
        for c in df.columns:
            schema.setdefault(str(c), []).append(sn)
    all_names = sorted(schema.keys())

    results = []
    n_ok = 0
    n_miss = 0
    for req in requested:
        req_s = str(req)
        if req_s in schema:
            results.append({
                "requested": req_s,
                "status": "ok",
                "found_in_sheets": schema[req_s],
            })
            n_ok += 1
            continue
        # 不存在 → 多策略找最近似
        ratios = []
        for name in all_names:
            r, strategy = _col_similarity(req_s, name)
            ratios.append((r, name, strategy))
        ratios.sort(key=lambda x: -x[0])
        top = ratios[:3]
        best_ratio, best_name, best_strategy = (top[0] if top else (0.0, "", "raw"))
        if best_ratio >= _COL_VALIDATE_MIN_RATIO:
            results.append({
                "requested": req_s,
                "status": "missing_with_suggestion",
                "best_match": best_name,
                "best_match_repr": repr(best_name),
                "similarity": round(best_ratio, 3),
                "match_strategy": best_strategy,
                "suggested_in_sheets": schema.get(best_name, []),
                "other_candidates": [
                    {"name": n, "similarity": round(r, 3), "strategy": s}
                    for r, n, s in top[1:] if r >= _COL_VALIDATE_MIN_RATIO
                ],
                "fix_hint": (
                    f"把脚本里的 df[{req_s!r}] 改成 df[{best_name!r}]"
                    + ("（识别为中→英别名幻觉）" if best_strategy == "cjk_substring" else "")
                ),
            })
        else:
            # 实在没近似 — 凭空捏造的高概率信号
            results.append({
                "requested": req_s,
                "status": "missing_no_suggestion",
                "best_match": best_name if best_name else None,
                "similarity": round(best_ratio, 3) if best_name else None,
                "note": (
                    "schema 中不存在此列且无近似匹配（最高相似度 "
                    f"{best_ratio:.0%}）— 你可能凭空想象了一个列名。"
                    "请重新查看 column_info 选用真实列名，或检查是否需要在分析中"
                    "新建该列（如 df['销售额'] = df['单价'] * df['数量']）"
                ),
            })
        n_miss += 1

    return {
        "ok": n_miss == 0,
        "file": str(path.name),
        "summary": {
            "total_requested": len(requested),
            "matched": n_ok,
            "missing": n_miss,
        },
        "schema_columns_count": len(all_names),
        "schema_sheets": list(sheets.keys()),
        "results": results,
        "rule": (
            "⚠️ 列名访问必须 Python 字面相等。任何空格、全/半角、大小写差异 → KeyError。"
            " 见 column_safety_card 或 column_name_quirk traps。"
        ),
    }


def analyze(filepath, sheet=None, do_stats=False, do_quality=False, as_json=False):
    path = Path(filepath)
    if not path.exists():
        print(f"Error: File not found: {filepath}", file=sys.stderr)
        cands = _fuzzy_find_candidates(filepath)
        if cands:
            print(
                f"Hint: 在 cwd 附近找到 {len(cands)} 个近似文件，"
                f"建议改用：{cands[0]['path']}（{cands[0]['reason']}）",
                file=sys.stderr,
            )
            if len(cands) > 1:
                for c in cands[1:]:
                    print(f"      备选：{c['path']}（{c['reason']}）", file=sys.stderr)
        sys.exit(1)
    sheets_data = read_file(filepath, sheet)
    output = {'file': str(path.name), 'sheets': {}}
    for sheet_name, df in sheets_data.items():
        sheet_info = {
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': list(df.columns),
        }
        if do_quality:
            sheet_info['quality'] = quality_audit(df)
        if do_stats:
            sheet_info['statistics'] = compute_stats(df)
        if not as_json:
            sheet_info['preview'] = format_preview(df)
        output['sheets'][sheet_name] = sheet_info
    if as_json:
        print(json.dumps(output, indent=2, default=str, ensure_ascii=False))
    else:
        print(f"\n{'='*60}\nFile: {path.name}\n{'='*60}")
        for sname, info in output['sheets'].items():
            print(f"\nSheet: {sname}")
            print(f"  Shape: {info['rows']} rows × {info['columns']} columns")
            print(f"  Columns: {', '.join(str(c) for c in info['column_names'])}")
            if do_quality:
                q = info['quality']
                if q['missing']:
                    print(f"  Missing values: {q['missing']}")
                if q['duplicate_rows']:
                    print(f"  Duplicate rows: {q['duplicate_rows']}")
                if q['empty_columns']:
                    print(f"  Empty columns: {q['empty_columns']}")
                if q['mixed_types']:
                    print(f"  Mixed types: {q['mixed_types']}")
                for issue in q['potential_issues']:
                    print(f"  ⚠ {issue}")
            if do_stats:
                print(f"\n  Statistics:")
                for col, s in info['statistics'].items():
                    print(f"    {col}: sum={s['sum']:,.2f}, mean={s['mean']:,.2f}, min={s['min']:,.2f}, max={s['max']:,.2f}")
            if 'preview' in info:
                print(f"\n  Preview (first 5 rows):")
                for line in info['preview'].split('\n'):
                    print(f"    {line}")


def main():
    parser = argparse.ArgumentParser(description='Read and analyze Excel/CSV files')
    parser.add_argument('file', help='Input file path')
    parser.add_argument('--inspect', action='store_true',
                        help='v4 INSPECT: full JSON report (workflow Step 1)')
    parser.add_argument('--validate-columns', nargs='+', metavar='COL',
                        dest='validate_columns',
                        help='预校验脚本将要用的列名是否存在；不存在则给 fuzzy 建议。'
                             '推荐在 Step 2 写完分析脚本后调用，挡掉列名幻觉/中→英别名/拼写错。')
    parser.add_argument('--merged-scan', action='store_true', dest='merged_scan',
                        help='按需检测合并单元格/多级表头（大文件流式 inspect 跳过了这步，'
                             '怀疑多级表头时用这个补检；流式低内存）。')
    parser.add_argument('--sheet', help='Specific sheet name (legacy)')
    parser.add_argument('--stats', action='store_true', help='Extended stats (legacy)')
    parser.add_argument('--quality', action='store_true', help='Quality audit (legacy)')
    parser.add_argument('--json', action='store_true', dest='as_json', help='JSON output (legacy)')
    args = parser.parse_args()

    if args.inspect:
        result = inspect_file(args.file)
        _v = _skill_version()
        if _v: result["skill_version"] = _v
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        sys.exit(0 if result.get("ok") else 1)

    if args.validate_columns:
        result = validate_columns(args.file, args.validate_columns)
        _v = _skill_version()
        if _v: result["skill_version"] = _v
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        sys.exit(0 if result.get("ok") else 1)

    if args.merged_scan:
        result = merged_scan(args.file)
        _v = _skill_version()
        if _v: result["skill_version"] = _v
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        sys.exit(0 if result.get("ok") else 1)

    analyze(args.file, sheet=args.sheet, do_stats=args.stats,
            do_quality=args.quality, as_json=args.as_json)


if __name__ == '__main__':
    main()
